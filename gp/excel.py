import requests
from bs4 import BeautifulSoup
import openpyxl

search_url = 'https://cn.investing.com/search/?q='

# 添加头部
headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.93 Safari/537.36'
}

# 打开Excel表格
wb = openpyxl.load_workbook(r'C:\Users\Admin\Desktop\abc.xlsx')
sheet = wb['Sheet1']

# 遍历企业名称列，获取相应的信息
for row in sheet.iter_rows(min_row=1, values_only=True):
    company_name = row[0]  # 假设企业名称在第一列

    # 搜索企业
    query_url = f'{search_url}{company_name}'
    print(query_url)
    response = requests.get(query_url, headers=headers)
    soup = BeautifulSoup(response.content, 'html.parser')

    # 获取跳转链接
    quote_link = soup.xpath('//a[@class="js-inner-all-results-quote-item row"]/@href')[0]
    quote_response = requests.get(quote_link, headers=headers)
    quote_soup = BeautifulSoup(quote_response.content, 'html.parser')

    profile_link = quote_soup.select('li')[-2]['href']
    profile_response = requests.get(profile_link, headers=headers)
    profile_soup = BeautifulSoup(profile_response.content, 'html.parser')

    # 获取企业简介
    profile_desc = profile_soup.select_one('#profile-fullStory-showhide').get_text(strip=True)

    # 获取企业地址
    address_spans = profile_soup.select('div.companyAddress span.float_lang_base_2.text_align_lang_base_2.dirLtr')
    address_texts = [span.get_text(strip=True) for span in address_spans]
    address = ''.join(address_texts)

    # 获取电话
    phone_span = profile_soup.select_one('div.companyPhone span.float_lang_base_2.text_align_lang_base_2.dirLtr')
    phone = phone_span.get_text(strip=True) if phone_span else ''

    # 获取网址
    web_span = profile_soup.select_one('div.companyWeb span.float_lang_base_2.text_align_lang_base_2.dirLtr')
    web = web_span.get_text(strip=True) if web_span else ''

    # 将获取到的信息添加到Excel表格对应的列中
    row_index = sheet.find(company_name).row  # 找到企业名称所在的行
    sheet.cell(row=row_index, column=11, value=profile_desc)  # 假设企业简介需要写入
    sheet.cell(row=row_index, column=10, value=address)  # 假设企业地址需要写入
    sheet.cell(row=row_index, column=7, value=phone)  # 假设电话需要写入
    sheet.cell(row=row_index, column=9, value=web)  # 假设网址需要写入

# 保存Excel表格
wb.save(r'C:\Users\Admin\Desktop\abc.xlsx')
